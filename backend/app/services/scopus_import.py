"""Import Scopus export file: mark existing articles and create missing ones.

Supports standard Scopus CSV/XLSX exports. Matching strategy:
  1. By DOI (exact, case-insensitive) — primary
  2. By title + year (normalized) — fallback for articles without DOI
  3. Create new articles for unmatched entries with source="scopus"
"""

import hashlib
import io
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date

import openpyxl
from sqlalchemy import func, select, update
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Article
from app.schemas.article import ArticleCreate
from app.utils.type_mapping import normalize_type

logger = logging.getLogger(__name__)

# Columns we look for (case-insensitive) in the uploaded file
_COL_NAMES: dict[str, set[str]] = {
    "doi": {"doi"},
    "title": {"title"},
    "year": {"year"},
    "link": {"link"},
    "authors": {"authors"},
    "authors_full": {"author full names"},
    "source_title": {"source title"},
    "volume": {"volume"},
    "issue": {"issue"},
    "page_start": {"page start"},
    "page_end": {"page end"},
    "cited_by": {"cited by"},
    "doc_type": {"document type"},
    "language": {"language of original document"},
    "open_access": {"open access"},
}

# Scopus link pattern to extract DOI if a DOI column is absent
_SCOPUS_DOI_RE = re.compile(r"10\.\d{4,9}/[^\s]+", re.IGNORECASE)
_AUTHOR_ID_RE = re.compile(r"\s*\(\d+\)")  # strip Scopus Author IDs like (59518974700)

# Scopus doc type -> our type mapping
_SCOPUS_TYPE_MAP: dict[str, str] = {
    "Article": "Article",
    "Conference paper": "Conference papers",
    "Conference Paper": "Conference papers",
    "Review": "Article",
    "Book chapter": "book-chapter",
    "Book Chapter": "book-chapter",
    "Book": "book",
    "Editorial": "Article",
    "Note": "Article",
    "Short survey": "Article",
    "Short Survey": "Article",
    "Letter": "Article",
    "Erratum": "Article",
    "Retracted": "Article",
}


def _normalize_title(title: str) -> str:
    """Lower-case, strip accents, collapse whitespace, remove punctuation."""
    title = unicodedata.normalize("NFKD", title)
    title = "".join(c for c in title if not unicodedata.combining(c))
    title = re.sub(r"[^\w\s]", "", title.lower())
    return re.sub(r"\s+", " ", title).strip()


def _find_col(headers: list[str], names: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h.strip().lower() in names:
            return i
    return None


def _scopus_num_id(title: str, year: int) -> str:
    """Generate a stable num_id for Scopus-sourced articles."""
    key = f"{title.strip().lower()}|{year}"
    return f"sc_{hashlib.md5(key.encode()).hexdigest()[:16]}"


@dataclass
class ScopusEntry:
    """A single parsed row from a Scopus export file."""
    title: str
    year: int | None = None
    doi: str | None = None
    authors: str = ""
    source_title: str = ""
    volume: str = ""
    issue: str = ""
    page_start: str = ""
    page_end: str = ""
    cited_by: int | None = None
    doc_type: str = ""
    language: str = ""
    link: str = ""
    norm_title: str = ""

    def __post_init__(self):
        self.norm_title = _normalize_title(self.title) if self.title else ""


def _parse_xlsx_full(file_bytes: bytes) -> list[ScopusEntry]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.worksheets[0]

    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip() if c else "" for c in row]

    cols = {key: _find_col(headers, names) for key, names in _COL_NAMES.items()}

    def _cell(row, key: str) -> str:
        idx = cols.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    entries: list[ScopusEntry] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = _cell(row, "title")
        if not title:
            continue

        # DOI
        doi = _cell(row, "doi").lower() or None
        if not doi:
            link = _cell(row, "link")
            m = _SCOPUS_DOI_RE.search(link)
            doi = m.group(0).lower() if m else None

        # Year
        year_raw = _cell(row, "year")
        try:
            year_val = int(year_raw) if year_raw else None
        except (ValueError, TypeError):
            year_val = None

        # Authors — prefer full names (strip Scopus IDs), fallback to short
        authors = _cell(row, "authors_full") or _cell(row, "authors")
        authors = _AUTHOR_ID_RE.sub("", authors)

        # Cited by
        cited_raw = _cell(row, "cited_by")
        try:
            cited_by = int(cited_raw) if cited_raw else None
        except (ValueError, TypeError):
            cited_by = None

        entries.append(ScopusEntry(
            title=title,
            year=year_val,
            doi=doi,
            authors=authors,
            source_title=_cell(row, "source_title"),
            volume=_cell(row, "volume"),
            issue=_cell(row, "issue"),
            page_start=_cell(row, "page_start"),
            page_end=_cell(row, "page_end"),
            cited_by=cited_by,
            doc_type=_cell(row, "doc_type"),
            language=_cell(row, "language"),
            link=_cell(row, "link"),
        ))

    wb.close()
    return entries


def _parse_csv_full(file_bytes: bytes) -> list[ScopusEntry]:
    import csv

    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    field_map = {f.strip().lower(): f for f in fieldnames}

    def _get(row_data: dict, key: str) -> str:
        for name in _COL_NAMES.get(key, set()):
            actual = field_map.get(name)
            if actual and row_data.get(actual):
                return row_data[actual].strip()
        return ""

    entries: list[ScopusEntry] = []
    for row_data in reader:
        title = _get(row_data, "title")
        if not title:
            continue

        doi = _get(row_data, "doi").lower() or None
        if not doi:
            link = _get(row_data, "link")
            m = _SCOPUS_DOI_RE.search(link)
            doi = m.group(0).lower() if m else None

        year_raw = _get(row_data, "year")
        try:
            year_val = int(year_raw) if year_raw else None
        except (ValueError, TypeError):
            year_val = None

        authors = _get(row_data, "authors_full") or _get(row_data, "authors")
        authors = _AUTHOR_ID_RE.sub("", authors)

        cited_raw = _get(row_data, "cited_by")
        try:
            cited_by = int(cited_raw) if cited_raw else None
        except (ValueError, TypeError):
            cited_by = None

        entries.append(ScopusEntry(
            title=title,
            year=year_val,
            doi=doi,
            authors=authors,
            source_title=_get(row_data, "source_title"),
            volume=_get(row_data, "volume"),
            issue=_get(row_data, "issue"),
            page_start=_get(row_data, "page_start"),
            page_end=_get(row_data, "page_end"),
            cited_by=cited_by,
            doc_type=_get(row_data, "doc_type"),
            language=_get(row_data, "language"),
            link=_get(row_data, "link"),
        ))

    return entries


def _entry_to_article_create(entry: ScopusEntry) -> ArticleCreate:
    """Convert a ScopusEntry to ArticleCreate for DB insertion."""
    published_at = None
    if entry.year:
        published_at = date(entry.year, 1, 1)

    raw_type = _SCOPUS_TYPE_MAP.get(entry.doc_type, entry.doc_type or None)

    return ArticleCreate(
        num_id=_scopus_num_id(entry.title, entry.year or 0),
        title=entry.title,
        authors=entry.authors or "Unknown",
        doi=entry.doi,
        published_at=published_at,
        journal_name=entry.source_title or None,
        type=normalize_type(raw_type),
        cited_by_count=entry.cited_by,
        language=entry.language[:10] if entry.language else None,
        source="scopus",
    )


async def mark_scopus_from_file(
    session: AsyncSession,
    file_bytes: bytes,
    filename: str,
) -> dict[str, int]:
    """Parse Scopus file, mark matching articles, and create missing ones.

    Returns dict with counts.
    """
    if filename.endswith(".csv"):
        entries = _parse_csv_full(file_bytes)
    else:
        entries = _parse_xlsx_full(file_bytes)

    logger.info("Scopus file parsed: %d entries", len(entries))

    # Build lookup indexes from the file
    doi_entries: dict[str, ScopusEntry] = {}
    title_year_entries: dict[tuple[str, int], ScopusEntry] = {}
    for e in entries:
        if e.doi:
            doi_entries[e.doi] = e
        if e.norm_title and e.year:
            title_year_entries[(e.norm_title, e.year)] = e

    matched_by_doi = 0
    matched_by_title = 0
    matched_entry_keys: set[int] = set()  # indices of matched entries

    # Step 1: Match by DOI — mark existing articles as in_scopus
    if doi_entries:
        doi_list = list(doi_entries.keys())
        for i in range(0, len(doi_list), 500):
            chunk = doi_list[i : i + 500]
            stmt = (
                update(Article)
                .where(func.lower(Article.doi).in_(chunk))
                .where(Article.in_scopus == False)  # noqa: E712
                .values(in_scopus=True)
            )
            result = await session.execute(stmt)
            matched_by_doi += result.rowcount

        # Find which entries were matched to exclude from creation
        matched_dois = set()
        for i in range(0, len(doi_list), 500):
            chunk = doi_list[i : i + 500]
            rows = (await session.execute(
                select(func.lower(Article.doi)).where(func.lower(Article.doi).in_(chunk))
            )).scalars().all()
            matched_dois.update(rows)

        for idx, e in enumerate(entries):
            if e.doi and e.doi in matched_dois:
                matched_entry_keys.add(idx)

    # Step 2: Match by title + year — mark existing articles
    # Load all non-matched articles from DB
    stmt = select(
        Article.id,
        Article.title,
        func.extract("year", Article.published_at).label("year"),
    ).where(
        Article.in_scopus == False,  # noqa: E712
        Article.published_at.is_not(None),
    )
    db_rows = (await session.execute(stmt)).all()

    db_matched_keys: set[tuple[str, int]] = set()
    ids_to_mark: list[int] = []
    for row in db_rows:
        norm = _normalize_title(row.title)
        year_val = int(row.year)
        key = (norm, year_val)
        if key in title_year_entries:
            ids_to_mark.append(row.id)
            db_matched_keys.add(key)

    if ids_to_mark:
        for i in range(0, len(ids_to_mark), 500):
            chunk = ids_to_mark[i : i + 500]
            await session.execute(
                update(Article).where(Article.id.in_(chunk)).values(in_scopus=True)
            )
        matched_by_title = len(ids_to_mark)

    # Mark title-matched entries
    for idx, e in enumerate(entries):
        if e.norm_title and e.year and (e.norm_title, e.year) in db_matched_keys:
            matched_entry_keys.add(idx)

    # Step 3: Create new articles for unmatched entries
    new_articles: list[ArticleCreate] = []
    for idx, e in enumerate(entries):
        if idx in matched_entry_keys:
            continue
        if not e.title:
            continue
        new_articles.append(_entry_to_article_create(e))

    created = 0
    if new_articles:
        # Split by DOI presence for proper conflict handling
        with_doi = [a for a in new_articles if a.doi]
        without_doi = [a for a in new_articles if not a.doi]

        if with_doi:
            values = [a.model_dump() for a in with_doi]
            stmt = (
                insert(Article)
                .values(values)
                .on_conflict_do_nothing(index_elements=["doi"])
            )
            result = await session.execute(stmt)
            created += result.rowcount

        if without_doi:
            values = [a.model_dump() for a in without_doi]
            stmt = (
                insert(Article)
                .values(values)
                .on_conflict_do_nothing(index_elements=["num_id"])
            )
            result = await session.execute(stmt)
            created += result.rowcount

        # Mark all newly created articles as in_scopus
        if created:
            await session.execute(
                update(Article)
                .where(Article.source == "scopus")
                .where(Article.in_scopus == False)  # noqa: E712
                .values(in_scopus=True)
            )

    await session.commit()

    # Count total marked
    total_marked = (
        await session.execute(
            select(func.count(Article.id)).where(Article.in_scopus == True)  # noqa: E712
        )
    ).scalar() or 0

    return {
        "total_in_file": len(entries),
        "matched_by_doi": matched_by_doi,
        "matched_by_title": matched_by_title,
        "created_new": created,
        "total_scopus": total_marked,
    }
